import openpyxl
import os
from pom_pytest_model.data.object_path import excel_as


class ExcelUtil:

    def read_1(self):
        wb = openpyxl.load_workbook(excel_as)
        sheet = wb['1']
        all_list = []
        for rows in range(2, sheet.max_row+1):
            temp_list = []
            for cols in range(1, sheet.max_column+1):
                temp_list.append(sheet.cell(rows, cols).value)
            all_list.append(temp_list)
        return all_list

    def read_2(self):
        wb = openpyxl.load_workbook(excel_as)
        sheet = wb['2']
        all_list = []
        for rows in range(2, sheet.max_row+1):
            temp_list = []
            for cols in range(2, sheet.max_column+1):
                temp_list.append(sheet.cell(rows, cols).value)
            all_list.append(temp_list)
        return all_list


if __name__ == '__main__':
    ExcelUtil().read_1()
    ExcelUtil().read_2()